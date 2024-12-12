# 6월 25일
# case1. fully funded swap은 원금을 주면 거래 상대방이 상환 시 원금과 이자를 증권사에게 지급한다.
# case2. unfunded swap 
# case3. Hi-five

# 엑셀 파일에서 case 별로 count 한다.
## ========function==============##

### 1. search 2. create 3. clear 4. update 5. append
import datetime
import pandas as pd
import numpy as np
import openpyxl as xl
import win32com.client
import time
from datetime import date,timedelta
from openpyxl.utils import get_column_letter as get_letter


class FX_data() :
    def __init__(self, *args) :
        
        #instance
        ## date, folderName, path
        self.foldername = ""
        self.downpath = ""
        self.input_date = ""

        #main  부분에 대한 statement 기술
        pass

    def __delattr__(self, name: str) -> None:
        pass

    #####Get Mail's attachment from outlook
    def att_down(self, foldername, downpath, input_date):
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI") 
        inbox = outlook.GetDefaultFolder(6).Folders['NHAR']
        messages = inbox.Items

        #today는 main 에서 설정하는 local 변수
        for ms in messages:
                print(ms.Subject)
                if ("NHAR_해외" in ms.Subject) and (str(ms.ReceivedTime)[0:10].replace("-","") > self.input_date):
                    attachments = ms.Attachments 
                    r = attachments.count
                    print("해외 메일 첨부파일 개수 : ",str(r)+"개") 
                    for i in range(1, r + 1): 
                        attachment = attachments.Item(i) 
                        attachment.SaveASFile(self.downpath +"해외"+str(attachment)) 
                    pass

    def scan_row(self, read_path,read_sheet) :
        self. read_path = read_path
        self. read_sheet = read_sheet

        xl_path             =   xl.load_workbook(self.read_path)
        target_sheet        =   xl_path['{}'.format(self.read_sheet)]
        max_row             =   len(target_sheet['A'])
        return int(max_row)


    #####Clear Sheet######
    def clear_sheet(self, file_path, sheet, start_cell, end_cell, return_path):
        #!!!!!!!!!!!!빈파일 에러 처리!!!!!!!!!!!!!!!!
        self.file_path = file_path
        self.sheet = sheet
        self.start_cell = start_cell
        self.end_cell = end_cell
        self.return_path = return_path

        try:
            xl_path         =   xl.load_workbook(self.file_path)                                         
            target_sheet    =   xl_path['{}'.format(self.sheet)]                                        
            tmp_rng         =   target_sheet['{}:{}'.format(self.start_cell,self.end_cell)]                   

            for x in tmp_rng:                                                                      
                for i in x:
                    if x[0].value is None:
                        break
                    i.value=""
                
            xl_path.save(self.return_path)                                                              
            print(self.sheet + "  sheet clear!!" )

        except Exception as e:
            print(e)
            exit(0)
        

    ######Copy data and Paste to output######
    def copy_data(self, read_path,read_sheet,write_path,write_sheet,write_start,location=0):
        # update
        self.read_path = read_path
        self.sheet = read_sheet
        self.write_path = write_path
        self.write_sheet = write_sheet
        self.write_start = write_start
        try:
            #Get last data
            xl_path         =   xl.load_workbook(self.read_path)
            target_sheet    =   xl_path['{}'.format(self.read_sheet)]
            max_column      =   str(get_letter(target_sheet.max_column))
            max_row         =   str(target_sheet.max_row)
            
            xl_wpath        =   xl.load_workbook(self.write_path)
            target_wsheet   =   xl_wpath['{}'.format(self.write_sheet)]
            start_row       =   self.write_start + target_wsheet.max_row -1

            #Copy, Paste
            excel           =   win32com.client.Dispatch("Excel.Application")
            wb_read         =   excel.Workbooks.Open(self.read_path)
            wb_write        =   excel.Workbooks.Open(self.write_path)
            sheet_read      =   wb_read.Sheets(self.read_sheet)
            sheet_write     =   wb_write.Sheets(self.write_sheet)

            sheet_read.Range("A2:{}{}".format(max_column,max_row)).Copy(sheet_write.Range('A{}'.format( self.write_start if location==0 else start_row)))
            wb_read.Close(True)
            wb_write.Close(True)
            print(read_sheet + "  sheet done!!" )

        except Exception as e:
            print(e)

        return 0